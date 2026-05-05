"""
Excel report writer.
"""

from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..logger import get_logger
from ..utils.flatten_utils import ListHandlingStrategy, flatten_dict
from .base import ReportWriter


class ExcelReportWriter(ReportWriter):
    """Excel report writer."""

    def __init__(self) -> None:
        """Initializes the Excel report writer."""
        self.logger = get_logger()

    def write_report(self, report_data: Dict[str, Any], file_path: str) -> None:
        """
        Writes a report to an Excel file with two sheets: Metadata and Tasks.

        Args:
            report_data: Report data to write
            file_path: Path to the report file
        """
        try:
            wb = Workbook()

            wb.remove(wb.active)

            # Create sheet for metadata
            metadata_sheet = wb.create_sheet("Metadata")
            self._write_metadata_sheet(metadata_sheet, report_data.get("metadata", {}))

            # Create sheet for tasks
            tasks_sheet = wb.create_sheet("Tasks")
            self._write_tasks_sheet(tasks_sheet, report_data.get("tasks", []))

            # Save file
            wb.save(file_path)
        except Exception as e:
            self.logger.error(
                f"Error writing Excel report to file {file_path}: {str(e)}"
            )
            raise

    def _write_metadata_sheet(self, sheet, metadata: Dict[str, Any]) -> None:
        """
        Writes metadata to sheet in key-value format.

        Args:
            sheet: Excel sheet to write to
            metadata: Metadata to write
        """
        if not metadata:
            return

        # Convert metadata to flat format
        flattened_metadata = flatten_dict(
            metadata, list_handling=ListHandlingStrategy.JSON_SERIALIZATION
        )

        row = 1
        for key, value in flattened_metadata.items():
            sheet[f"A{row}"] = key
            sheet[f"B{row}"] = str(value) if value is not None else ""
            row += 1

        # Automatically adjust column widths
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 50

    def _write_tasks_sheet(self, sheet, tasks: List[Dict[str, Any]]) -> None:
        """
        Writes tasks in tabular format.

        Args:
            sheet: Excel sheet to write to
            tasks: List of tasks to write
        """
        if not tasks:
            return

        # Convert all tasks to flat format
        flattened_tasks = [
            flatten_dict(task, list_handling=ListHandlingStrategy.JSON_SERIALIZATION)
            for task in tasks
        ]

        # Collect all unique keys for headers
        all_keys = set()
        for task in flattened_tasks:
            all_keys.update(task.keys())
        headers = sorted(all_keys)

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            sheet[f"{get_column_letter(col_idx)}1"] = header

        # Write task data
        for row_idx, task in enumerate(flattened_tasks, 2):
            for col_idx, header in enumerate(headers, 1):
                value = task.get(header, "")
                sheet[f"{get_column_letter(col_idx)}{row_idx}"] = (
                    str(value) if value is not None else ""
                )

        # Automatically adjust column widths
        for col_idx, header in enumerate(headers, 1):
            column_width = max(len(header), 15)  # Minimum width 15
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(
                column_width, 50
            )

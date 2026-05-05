"""
Loader for issues from Excel files.
"""

from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from ..utils.dot_notation_utils import convert_dot_notation_to_nested_dict
from .base_loader import BaseFileLoader
from .exception_handler import handle_loader_exceptions


class ExcelFileLoader(BaseFileLoader):
    """Loader for issues from Excel files with dot notation support in headers."""

    def _get_headers(self, sheet) -> List[str]:
        """Gets headers from the first row of the sheet."""
        headers = []
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header is not None:
                headers.append(str(header))
            else:
                headers.append(f"column_{col}")
        return headers

    def _read_row_data(
        self, sheet, row: int, headers: List[str]
    ) -> Tuple[Dict[str, str], bool]:
        """Reads data from a sheet row."""
        row_data = {}
        has_data = False

        for col_idx, header in enumerate(headers, 1):
            cell_value = sheet.cell(row=row, column=col_idx).value
            if cell_value is not None:
                row_data[header] = str(cell_value)
                has_data = True

        return row_data, has_data

    def _process_row_data(self, row_data: Dict[str, str]) -> Dict[Any, Any]:
        """Processes row data."""
        # Filter out empty values
        filtered_row = {k: v for k, v in row_data.items() if v is not None and v != ""}

        # Convert flat dictionary with dot notation to nested dictionary
        nested_dict = convert_dot_notation_to_nested_dict(filtered_row)

        # Process linking information if present
        if isinstance(nested_dict, dict) and "linking" in nested_dict:
            linking_info = nested_dict.pop("linking")
            if isinstance(linking_info, (dict, list)):
                if isinstance(linking_info, dict) and all(
                    isinstance(k, str) and k.isdigit() for k in linking_info.keys()
                ):
                    sorted_items = sorted(linking_info.items(), key=lambda x: int(x[0]))
                    linking_info = [v for _, v in sorted_items]
                # Add linking information directly to the issue
                nested_dict["linking"] = linking_info

        return nested_dict

    @handle_loader_exceptions(format_name="Excel file")
    def load_issues(self, file_path: str) -> List[Dict[Any, Any]]:
        """Loads issues from Excel file with dot notation support in headers."""
        # Validate file path (only validation, don't open file)
        # This is necessary to prevent directory traversal attacks
        self._validate_and_open_file(file_path, "r").close()

        # Load workbook
        workbook = load_workbook(str(file_path), read_only=True, data_only=True)

        # Work with active sheet
        sheet = workbook.active

        # Check that sheet is not empty
        if sheet.max_row == 0 or sheet.max_column == 0:
            raise ValueError("Excel file is empty or corrupted")

        # Get headers from first row
        headers = self._get_headers(sheet)

        # Check that headers exist
        if not headers:
            raise ValueError("Excel file does not contain headers")

        issues = []

        # Read data from remaining rows
        for row in range(2, sheet.max_row + 1):
            row_data, has_data = self._read_row_data(sheet, row, headers)

            # Skip empty rows
            if not has_data:
                continue

            # Process row data
            nested_dict = self._process_row_data(row_data)

            issues.append(nested_dict)

        workbook.close()
        return issues

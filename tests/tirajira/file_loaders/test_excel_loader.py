from unittest.mock import MagicMock, patch

import pytest

from tirajira.file_loaders.excel_loader import ExcelFileLoader


def test_excel_file_loader_success():
    """Тест: успешная загрузка задач из Excel файла"""
    # Создаем mock для workbook и sheet
    mock_sheet = MagicMock()
    mock_sheet.max_row = 3
    mock_sheet.max_column = 7

    # Настраиваем mock для ячеек с заголовками
    mock_sheet.cell.return_value.value = None
    mock_sheet.cell.side_effect = lambda row, column: MagicMock(
        value={
            (1, 1): "project.key",
            (1, 2): "summary",
            (1, 3): "description",
            (1, 4): "issuetype.name",
            (1, 5): "assignee.emailAddress",
            (1, 6): "priority.name",
            (1, 7): "epic_key",
            (2, 1): "PROJ",
            (2, 2): "Test task 1",
            (2, 3): "Description 1",
            (2, 4): "Task",
            (2, 5): "developer@example.com",
            (2, 6): "High",
            (2, 7): "PROJ-100",
            (3, 1): "PROJ",
            (3, 2): "Test task 2",
            (3, 3): "Description 2",
            (3, 4): "Bug",
            (3, 5): "qa@example.com",
            (3, 6): "Medium",
            (3, 7): "PROJ-101",
        }.get((row, column), None)
    )

    mock_workbook = MagicMock()
    mock_workbook.active = mock_sheet
    mock_workbook.close = MagicMock()

    # Используем mock для симуляции чтения файла
    with patch("os.path.exists", return_value=True):
        with patch(
            "tirajira.file_loaders.excel_loader.load_workbook",
            return_value=mock_workbook,
        ):
            loader = ExcelFileLoader()
            issues = loader.load_issues("test.xlsx")

            # Проверяем, что данные загрузились правильно
            assert len(issues) == 2
            assert issues[0]["summary"] == "Test task 1"
            assert issues[0]["project"]["key"] == "PROJ"
            assert issues[0]["issuetype"]["name"] == "Task"
            assert issues[0]["assignee"]["emailAddress"] == "developer@example.com"
            assert issues[0]["priority"]["name"] == "High"
            assert issues[0]["epic_key"] == "PROJ-100"

            assert issues[1]["summary"] == "Test task 2"
            assert issues[1]["issuetype"]["name"] == "Bug"
            assert issues[1]["assignee"]["emailAddress"] == "qa@example.com"
            assert issues[1]["priority"]["name"] == "Medium"
            assert issues[1]["epic_key"] == "PROJ-101"


def test_excel_file_loader_file_not_found():
    """Тест: обработка ошибки файла не найден"""
    with patch("os.path.exists", return_value=False):
        loader = ExcelFileLoader()
        with pytest.raises(FileNotFoundError):
            loader.load_issues("nonexistent.xlsx")


def test_excel_file_loader_invalid_excel():
    """Тест: обработка ошибки невалидного Excel файла"""
    with patch("os.path.exists", return_value=True):
        with patch(
            "tirajira.file_loaders.excel_loader.load_workbook"
        ) as mock_load_workbook:
            from openpyxl.utils.exceptions import InvalidFileException

            mock_load_workbook.side_effect = InvalidFileException(
                "Неверный формат Excel файла"
            )

            loader = ExcelFileLoader()
            with pytest.raises(ValueError, match="Ошибка парсинга Excel файла"):
                loader.load_issues("test.xlsx")


def test_excel_file_loader_empty_file():
    """Тест: обработка ошибки пустого Excel файла"""
    # Создаем mock для пустого workbook
    mock_sheet = MagicMock()
    mock_sheet.max_row = 0
    mock_sheet.max_column = 0

    mock_workbook = MagicMock()
    mock_workbook.active = mock_sheet
    mock_workbook.close = MagicMock()

    with patch("os.path.exists", return_value=True):
        with patch(
            "tirajira.file_loaders.excel_loader.load_workbook",
            return_value=mock_workbook,
        ):
            loader = ExcelFileLoader()
            with pytest.raises(ValueError, match="Excel файл пуст или поврежден"):
                loader.load_issues("test.xlsx")


def test_excel_file_loader_missing_headers():
    """Тест: обработка Excel файла с отсутствующими заголовками"""
    # Создаем mock для workbook с заголовками None
    mock_sheet = MagicMock()
    mock_sheet.max_row = 2
    mock_sheet.max_column = 2

    # Настраиваем mock для ячеек с None заголовками
    mock_sheet.cell.return_value.value = None
    mock_sheet.cell.side_effect = lambda row, column: MagicMock(
        value=None if row == 1 else f"value_{row}_{column}"
    )

    mock_workbook = MagicMock()
    mock_workbook.active = mock_sheet
    mock_workbook.close = MagicMock()

    with patch("os.path.exists", return_value=True):
        with patch(
            "tirajira.file_loaders.excel_loader.load_workbook",
            return_value=mock_workbook,
        ):
            loader = ExcelFileLoader()
            issues = loader.load_issues("test.xlsx")

            # Проверяем, что данные загрузились с автоматически
            # сгенерированными заголовками
            assert len(issues) == 1
            assert issues[0]["column_1"] == "value_2_1"
            assert issues[0]["column_2"] == "value_2_2"


def test_excel_file_loader_skip_empty_rows():
    """Тест: пропуск пустых строк в Excel файле"""
    # Создаем mock для workbook с пустыми строками
    mock_sheet = MagicMock()
    mock_sheet.max_row = 4
    mock_sheet.max_column = 2

    # Настраиваем mock для ячеек с пустыми строками
    mock_sheet.cell.return_value.value = None
    mock_sheet.cell.side_effect = lambda row, column: MagicMock(
        value={
            (1, 1): "project.key",
            (1, 2): "summary",
            (2, 1): "PROJ",
            (2, 2): "Task 1",
            (3, 1): None,  # Пустая строка
            (3, 2): None,  # Пустая строка
            (4, 1): "PROJ",
            (4, 2): "Task 2",
        }.get((row, column), None)
    )

    mock_workbook = MagicMock()
    mock_workbook.active = mock_sheet
    mock_workbook.close = MagicMock()

    with patch("os.path.exists", return_value=True):
        with patch(
            "tirajira.file_loaders.excel_loader.load_workbook",
            return_value=mock_workbook,
        ):
            loader = ExcelFileLoader()
            issues = loader.load_issues("test.xlsx")

            # Проверяем, что пустые строки были пропущены
            assert len(issues) == 2
            assert issues[0]["project"]["key"] == "PROJ"
            assert issues[0]["summary"] == "Task 1"
            assert issues[1]["project"]["key"] == "PROJ"
            assert issues[1]["summary"] == "Task 2"


def test_excel_file_loader_general_exception():
    """Тест: обработка общих исключений при чтении файла."""
    with patch("os.path.exists", return_value=True):
        with patch(
            "tirajira.file_loaders.excel_loader.load_workbook",
            side_effect=Exception("Ошибка файловой системы"),
        ):
            loader = ExcelFileLoader()
            with pytest.raises(
                Exception, match="Ошибка при чтении файла: Ошибка файловой системы"
            ):
                loader.load_issues("test.xlsx")


def test_excel_file_loader_nested_dict_conversion():
    """Тест: преобразование вложенных словарей с точечной нотацией."""
    # Создаем mock для workbook с вложенными ключами
    mock_sheet = MagicMock()
    mock_sheet.max_row = 2
    mock_sheet.max_column = 4

    # Настраиваем mock для ячеек с вложенными ключами
    mock_sheet.cell.return_value.value = None
    mock_sheet.cell.side_effect = lambda row, column: MagicMock(
        value={
            (1, 1): "project.key",
            (1, 2): "issuetype.name",
            (1, 3): "assignee.emailAddress",
            (1, 4): "simple_key",
            (2, 1): "PROJ",
            (2, 2): "Task",
            (2, 3): "user@example.com",
            (2, 4): "value",
        }.get((row, column), None)
    )

    mock_workbook = MagicMock()
    mock_workbook.active = mock_sheet
    mock_workbook.close = MagicMock()

    with patch("os.path.exists", return_value=True):
        with patch(
            "tirajira.file_loaders.excel_loader.load_workbook",
            return_value=mock_workbook,
        ):
            loader = ExcelFileLoader()
            issues = loader.load_issues("test.xlsx")

            # Проверяем, что данные загрузились правильно с вложенными словарями
            assert len(issues) == 1
            assert issues[0]["project"]["key"] == "PROJ"
            assert issues[0]["issuetype"]["name"] == "Task"
            assert issues[0]["assignee"]["emailAddress"] == "user@example.com"
            assert issues[0]["simple_key"] == "value"

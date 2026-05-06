"""
Consolidated parametrized tests for all file loaders.
"""

import csv
import json
from unittest.mock import MagicMock, Mock, mock_open, patch

import pytest
import yaml
from openpyxl.utils.exceptions import InvalidFileException

from tirajira.file_loaders.csv_loader import CsvFileLoader
from tirajira.file_loaders.excel_loader import ExcelFileLoader
from tirajira.file_loaders.json_loader import JsonFileLoader
from tirajira.file_loaders.xml_loader import XmlFileLoader
from tirajira.file_loaders.yaml_loader import YamlFileLoader


class TestAllFileLoaders:
    """Parametrized tests for all file loader implementations."""

    # Test data for common success case
    COMMON_TEST_DATA = [
        {
            "project": {"key": "PROJ"},
            "summary": "Test task 1",
            "issuetype": {"name": "Task"},
        },
        {
            "project": {"key": "PROJ"},
            "summary": "Test task 2",
            "issuetype": {"name": "Bug"},
        },
    ]

    # Parametrized test cases for different file formats
    # Each tuple contains: (loader_class, file_extension, test_data_converter,
    # additional_test_cases)
    @pytest.mark.parametrize(
        "loader_class,file_extension,data_converter,format_specific_tests",
        [
            (
                JsonFileLoader,
                "json",
                lambda data: json.dumps(data, ensure_ascii=False),
                [
                    ("test_not_list", "json_not_list_test"),
                    ("test_invalid_format", "json_invalid_test"),
                ],
            ),
            (
                YamlFileLoader,
                "yaml",
                lambda data: yaml.dump(data, allow_unicode=True),
                [
                    ("test_not_list", "yaml_not_list_test"),
                    ("test_invalid_format", "yaml_invalid_test"),
                ],
            ),
            (
                CsvFileLoader,
                "csv",
                lambda data: TestAllFileLoaders._convert_to_csv(data),
                [
                    ("test_empty_file", "csv_empty_test"),
                    ("test_skip_empty_rows", "csv_skip_empty_rows_test"),
                    ("test_nested_dict_conversion", "csv_nested_dict_test"),
                    ("test_invalid_format", "csv_invalid_test"),
                ],
            ),
            (
                ExcelFileLoader,
                "xlsx",
                lambda data: data,  # Excel uses mock objects instead of string content
                [
                    ("test_empty_file", "excel_empty_test"),
                    ("test_missing_headers", "excel_missing_headers_test"),
                    ("test_skip_empty_rows", "excel_skip_empty_rows_test"),
                    ("test_nested_dict_conversion", "excel_nested_dict_test"),
                    ("test_invalid_format", "excel_invalid_test"),
                ],
            ),
            (
                XmlFileLoader,
                "xml",
                lambda data: TestAllFileLoaders._convert_to_xml(data),
                [
                    ("test_invalid_format", "xml_invalid_test"),
                    ("test_wrong_root_element", "xml_wrong_root_test"),
                    ("test_complex_structure", "xml_complex_test"),
                    ("test_dot_notation", "xml_dot_notation_test"),
                    ("test_merge_dicts", "xml_merge_dicts_test"),
                    ("test_element_to_dict_text_content", "xml_text_content_test"),
                    ("test_element_to_dict_mixed_content", "xml_mixed_content_test"),
                    ("test_element_to_dict_list_elements", "xml_list_elements_test"),
                    ("test_custom_fields_processing", "xml_custom_fields_test"),
                    ("test_nested_elements_with_attributes", "xml_nested_attrs_test"),
                    ("test_mixed_content_with_tail", "xml_mixed_tail_test"),
                    ("test_empty_and_whitespace_elements", "xml_empty_whitespace_test"),
                    (
                        "test_dot_notation_in_element_names",
                        "xml_dot_notation_elements_test",
                    ),
                    ("test_merge_dicts_logic", "xml_merge_dicts_logic_test"),
                    ("test_process_list_field_logic", "xml_process_list_field_test"),
                    ("test_process_labels_logic", "xml_process_labels_test"),
                    ("test_process_attachments_logic", "xml_process_attachments_test"),
                    (
                        "test_process_customfields_post_logic",
                        "xml_process_customfields_post_test",
                    ),
                    ("test_process_attributes_logic", "xml_process_attributes_test"),
                    ("test_process_children_logic", "xml_process_children_test"),
                    (
                        "test_process_customfields_logic",
                        "xml_process_customfields_test",
                    ),
                    (
                        "test_process_child_element_logic",
                        "xml_process_child_element_test",
                    ),
                    (
                        "test_add_to_children_dict_logic",
                        "xml_add_to_children_dict_test",
                    ),
                    ("test_post_process_result_logic", "xml_post_process_result_test"),
                ],
            ),
        ],
    )
    def test_file_loader_success(
        self, loader_class, file_extension, data_converter, format_specific_tests
    ):
        """Test: successful loading of tasks from file"""
        # Prepare test data
        test_data = self.COMMON_TEST_DATA

        # Convert test data to format-specific representation
        formatted_data = data_converter(test_data)

        # Create loader instance
        loader = loader_class()

        # Mock file operations based on format
        if file_extension in ["json", "yaml", "csv", "xml"]:
            self._test_text_based_loaders(
                loader, file_extension, formatted_data, test_data
            )
        elif file_extension == "xlsx":
            self._test_excel_loader(loader, file_extension, test_data)

    def _test_text_based_loaders(
        self, loader, file_extension, formatted_data, test_data
    ):
        # For text-based formats, mock file content
        with patch("builtins.open", mock_open(read_data=formatted_data)):
            with patch("os.path.exists", return_value=True):
                # Additional mocks for CSV
                if file_extension == "csv":
                    self._test_csv_loader(
                        loader, file_extension, formatted_data, test_data
                    )
                # Additional mocks for XML
                elif file_extension == "xml":
                    issues = loader.load_issues(f"test.{file_extension}")
                    self._verify_loaded_data(issues, test_data)
                # Simple case for JSON and YAML
                else:
                    issues = loader.load_issues(f"test.{file_extension}")
                    self._verify_loaded_data(issues, test_data)

    def _test_csv_loader(self, loader, file_extension, formatted_data, test_data):
        # Parse the CSV content to determine fieldnames
        if formatted_data:
            csv_lines = formatted_data.strip().split("\n")
            fieldnames = csv_lines[0].split(",") if csv_lines else []
        else:
            fieldnames = []

        mock_reader = MagicMock()
        mock_reader.fieldnames = fieldnames

        # Convert CSV rows to list for iteration
        rows = []
        for issue in test_data:
            row = {}
            for key, value in issue.items():
                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        row[f"{key}.{sub_key}"] = sub_value
                else:
                    row[key] = value
            rows.append(row)

        mock_reader.__iter__ = Mock(return_value=iter(rows))
        with patch("csv.DictReader", return_value=mock_reader):
            with patch("csv.Sniffer.sniff") as mock_sniffer:
                mock_dialect = csv.excel()
                mock_dialect.delimiter = ","
                mock_sniffer.return_value = mock_dialect
                issues = loader.load_issues(f"test.{file_extension}")
                self._verify_loaded_data(issues, test_data)

    def _test_excel_loader(self, loader, file_extension, test_data):
        # For Excel, we need more complex mocking
        mock_sheet = MagicMock()
        mock_sheet.max_row = len(test_data) + 1  # +1 for header row
        mock_sheet.max_column = 3  # We'll use 3 columns for simplicity

        # Setup cell values
        cell_values = {
            (1, 1): "project.key",
            (1, 2): "summary",
            (1, 3): "issuetype.name",
        }

        # Add data rows
        for i, issue in enumerate(test_data):
            row_idx = i + 2  # +1 for 1-based indexing, +1 for header row
            cell_values[(row_idx, 1)] = issue["project"]["key"]
            cell_values[(row_idx, 2)] = issue["summary"]
            cell_values[(row_idx, 3)] = issue["issuetype"]["name"]

        mock_sheet.cell.side_effect = lambda row, column: MagicMock(
            value=cell_values.get((row, column), None)
        )

        mock_workbook = MagicMock()
        mock_workbook.active = mock_sheet
        mock_workbook.close = MagicMock()

        with patch("os.path.exists", return_value=True):
            with patch(
                "tirajira.file_loaders.excel_loader.load_workbook",
                return_value=mock_workbook,
            ):
                with patch.object(ExcelFileLoader, "_validate_and_open_file"):
                    issues = loader.load_issues(f"test.{file_extension}")
                    self._verify_loaded_data(issues, test_data)

    def _verify_loaded_data(self, issues, test_data):
        # Verify loaded data
        assert len(issues) == len(test_data)
        assert issues[0]["summary"] == "Test task 1"
        assert issues[1]["issuetype"]["name"] == "Bug"

    @pytest.mark.parametrize(
        "loader_class,file_extension,data_converter,format_specific_tests",
        [
            (
                JsonFileLoader,
                "json",
                lambda data: json.dumps(data, ensure_ascii=False, indent=2),
                [
                    ("test_simple_issue", "json_simple_test"),
                    ("test_complex_issue", "json_complex_test"),
                    ("test_nested_dict_conversion", "json_nested_dict_test"),
                    ("test_invalid_format", "json_invalid_test"),
                ],
            ),
            (
                YamlFileLoader,
                "yaml",
                lambda data: yaml.dump(
                    data, allow_unicode=True, default_flow_style=False
                ),
                [
                    ("test_simple_issue", "yaml_simple_test"),
                    ("test_complex_issue", "yaml_complex_test"),
                    ("test_nested_dict_conversion", "yaml_nested_dict_test"),
                    ("test_invalid_format", "yaml_invalid_test"),
                ],
            ),
            (
                CsvFileLoader,
                "csv",
                lambda data: TestAllFileLoaders._convert_to_csv(data),
                [
                    ("test_simple_issue", "csv_simple_test"),
                    ("test_complex_issue", "csv_complex_test"),
                    ("test_nested_dict_conversion", "csv_nested_dict_test"),
                    ("test_invalid_format", "csv_invalid_test"),
                ],
            ),
            (
                ExcelFileLoader,
                "xlsx",
                lambda data: TestAllFileLoaders._convert_to_excel(data),
                [
                    ("test_simple_issue", "excel_simple_test"),
                    ("test_complex_issue", "excel_complex_test"),
                    ("test_nested_dict_conversion", "excel_nested_dict_test"),
                    ("test_invalid_format", "excel_invalid_test"),
                ],
            ),
            (
                XmlFileLoader,
                "xml",
                lambda data: TestAllFileLoaders._convert_to_xml(data),
                [
                    ("test_invalid_format", "xml_invalid_test"),
                    ("test_wrong_root_element", "xml_wrong_root_test"),
                    ("test_complex_structure", "xml_complex_test"),
                    ("test_dot_notation", "xml_dot_notation_test"),
                    ("test_merge_dicts", "xml_merge_dicts_test"),
                    ("test_element_to_dict_text_content", "xml_text_content_test"),
                    ("test_element_to_dict_mixed_content", "xml_mixed_content_test"),
                    ("test_element_to_dict_list_elements", "xml_list_elements_test"),
                    ("test_custom_fields_processing", "xml_custom_fields_test"),
                    ("test_nested_elements_with_attributes", "xml_nested_attrs_test"),
                    ("test_mixed_content_with_tail", "xml_mixed_tail_test"),
                    ("test_empty_and_whitespace_elements", "xml_empty_whitespace_test"),
                    (
                        "test_dot_notation_in_element_names",
                        "xml_dot_notation_elements_test",
                    ),
                    ("test_merge_dicts_logic", "xml_merge_dicts_logic_test"),
                    ("test_process_list_field_logic", "xml_process_list_field_test"),
                    ("test_process_labels_logic", "xml_process_labels_test"),
                    ("test_process_attachments_logic", "xml_process_attachments_test"),
                    (
                        "test_process_customfields_post_logic",
                        "xml_process_customfields_post_test",
                    ),
                    ("test_process_attributes_logic", "xml_process_attributes_test"),
                    ("test_process_children_logic", "xml_process_children_test"),
                    (
                        "test_process_customfields_logic",
                        "xml_process_customfields_test",
                    ),
                    (
                        "test_process_child_element_logic",
                        "xml_process_child_element_test",
                    ),
                    (
                        "test_add_to_children_dict_logic",
                        "xml_add_to_children_dict_test",
                    ),
                    ("test_post_process_result_logic", "xml_post_process_result_test"),
                ],
            ),
        ],
    )
    def test_file_loader_file_not_found(
        self, loader_class, file_extension, data_converter, format_specific_tests
    ):
        """Test: handling file not found error"""
        with patch("os.path.exists", return_value=False):
            loader = loader_class()
            with pytest.raises(FileNotFoundError):
                loader.load_issues(f"nonexistent.{file_extension}")

    @pytest.mark.parametrize(
        "loader_class,file_extension,data_converter,format_specific_tests",
        [
            (
                JsonFileLoader,
                "json",
                lambda data: json.dumps(data, ensure_ascii=False),
                [
                    ("test_not_list", "json_not_list_test"),
                    ("test_invalid_format", "json_invalid_test"),
                ],
            ),
            (
                YamlFileLoader,
                "yaml",
                lambda data: yaml.dump(data, allow_unicode=True),
                [
                    ("test_not_list", "yaml_not_list_test"),
                    ("test_invalid_format", "yaml_invalid_test"),
                ],
            ),
            (
                CsvFileLoader,
                "csv",
                lambda data: TestAllFileLoaders._convert_to_csv(data),
                [
                    ("test_empty_file", "csv_empty_test"),
                    ("test_skip_empty_rows", "csv_skip_empty_rows_test"),
                    ("test_nested_dict_conversion", "csv_nested_dict_test"),
                    ("test_invalid_format", "csv_invalid_test"),
                ],
            ),
            (
                ExcelFileLoader,
                "xlsx",
                lambda data: data,
                [
                    ("test_empty_file", "excel_empty_test"),
                    ("test_missing_headers", "excel_missing_headers_test"),
                    ("test_skip_empty_rows", "excel_skip_empty_rows_test"),
                    ("test_nested_dict_conversion", "excel_nested_dict_test"),
                    ("test_invalid_format", "excel_invalid_test"),
                ],
            ),
            (
                XmlFileLoader,
                "xml",
                lambda data: TestAllFileLoaders._convert_to_xml(data),
                [
                    ("test_invalid_format", "xml_invalid_test"),
                    ("test_wrong_root_element", "xml_wrong_root_test"),
                    ("test_complex_structure", "xml_complex_test"),
                ],
            ),
        ],
    )
    def test_file_loader_general_exception(
        self, loader_class, file_extension, data_converter, format_specific_tests
    ):
        """Test: handling general exceptions when reading file"""
        loader = loader_class()

        if file_extension in ["json", "yaml", "csv", "xml"]:
            with patch("builtins.open", side_effect=Exception("File system error")):
                with patch("os.path.exists", return_value=True):
                    with pytest.raises(
                        Exception,
                        match="Error reading file: File system error",
                    ):
                        loader.load_issues(f"test.{file_extension}")
        elif file_extension == "xlsx":
            with patch("os.path.exists", return_value=True):
                with patch(
                    "tirajira.file_loaders.excel_loader.load_workbook",
                    side_effect=Exception("File system error"),
                ):
                    with patch.object(ExcelFileLoader, "_validate_and_open_file"):
                        with pytest.raises(
                            Exception,
                            match="Error reading file: File system error",
                        ):
                            loader.load_issues(f"test.{file_extension}")

    @staticmethod
    def _convert_to_csv(data):
        """Convert test data to CSV format."""
        if not data:
            return ""

        # Create header row
        headers = set()
        for issue in data:
            for key, value in issue.items():
                if isinstance(value, dict):
                    for sub_key in value.keys():
                        headers.add(f"{key}.{sub_key}")
                else:
                    headers.add(key)

        headers = sorted(headers)

        # Create rows
        rows = []
        for issue in data:
            row = []
            for header in headers:
                if "." in header:
                    main_key, sub_key = header.split(".", 1)
                    if (
                        main_key in issue
                        and isinstance(issue[main_key], dict)
                        and sub_key in issue[main_key]
                    ):
                        row.append(str(issue[main_key][sub_key]))
                    else:
                        row.append("")
                else:
                    row.append(str(issue.get(header, "")))
            rows.append(",".join(row))

        # Combine headers and rows
        csv_content = ",".join(headers) + "\n" + "\n".join(rows)
        return csv_content

    @staticmethod
    def _convert_to_xml(data):
        """Convert test data to XML format."""
        xml_parts = ['<?xml version="1.0" encoding="UTF-8"?><issues>']

        for issue in data:
            xml_parts.append("<issue>")
            for key, value in issue.items():
                if isinstance(value, dict):
                    # For nested dictionaries, create nested elements
                    xml_parts.append(f"<{key}>")
                    for sub_key, sub_value in value.items():
                        xml_parts.append(f"<{sub_key}>{sub_value}</{sub_key}>")
                    xml_parts.append(f"</{key}>")
                else:
                    # For simple values, create simple elements
                    xml_parts.append(f"<{key}>{value}</{key}>")
            xml_parts.append("</issue>")

        xml_parts.append("</issues>")
        return "".join(xml_parts)

    # Format-specific test cases
    def test_json_not_list(self):
        """Test: handling error when JSON does not contain an array"""
        # JSON with an object instead of an array
        json_content = '{"project": "PROJ", "summary": "Test task"}'

        with patch("builtins.open", mock_open(read_data=json_content)):
            with patch("os.path.exists", return_value=True):
                loader = JsonFileLoader()
                with pytest.raises(
                    ValueError, match="JSON file must contain an array of issues"
                ):
                    loader.load_issues("test.json")

    def test_json_invalid_format(self):
        """Test: handling invalid JSON error"""
        with patch("builtins.open", mock_open(read_data="invalid json")):
            with patch("os.path.exists", return_value=True):
                loader = JsonFileLoader()
                with pytest.raises(ValueError, match="Error parsing JSON file"):
                    loader.load_issues("test.json")

    def test_yaml_not_list(self):
        """Test: handling error when YAML does not contain an array"""
        # YAML with an object instead of an array
        yaml_content = "project: PROJ\nsummary: Test task"

        with patch("builtins.open", mock_open(read_data=yaml_content)):
            with patch("os.path.exists", return_value=True):
                loader = YamlFileLoader()
                with pytest.raises(
                    ValueError, match="YAML file must contain an array of issues"
                ):
                    loader.load_issues("test.yaml")

    def test_yaml_invalid_format(self):
        """Test: handling invalid YAML error"""
        with patch("builtins.open", mock_open(read_data="invalid: yaml: content:")):
            with patch("os.path.exists", return_value=True):
                loader = YamlFileLoader()
                with pytest.raises(ValueError, match="Error parsing YAML file"):
                    loader.load_issues("test.yaml")

    def test_csv_empty_file(self):
        """Test: handling empty CSV file"""
        # Mock DictReader with None fieldnames
        mock_reader = MagicMock()
        mock_reader.fieldnames = None

        with patch("builtins.open", mock_open(read_data="")):
            with patch("os.path.exists", return_value=True):
                with patch("csv.DictReader", return_value=mock_reader):
                    with patch("csv.Sniffer.sniff") as mock_sniffer:
                        # Configure mock for CSV dialect
                        mock_dialect = csv.excel()
                        mock_dialect.delimiter = ","
                        mock_sniffer.return_value = mock_dialect

                        loader = CsvFileLoader()
                        with pytest.raises(
                            Exception,
                            match=(
                                "Error reading file: CSV file is empty or corrupted"
                            ),
                        ):
                            loader.load_issues("empty.csv")

    def test_csv_skip_empty_rows(self):
        """Test: skipping empty rows in CSV file"""
        # Prepare test CSV data with empty rows
        csv_content = """project.key,summary
PROJ,Task 1

PROJ,Task 2"""

        # Create mock for DictReader with empty rows
        mock_reader = MagicMock()
        mock_reader.fieldnames = ["project.key", "summary"]
        mock_reader.__iter__ = Mock(
            return_value=iter(
                [
                    {"project.key": "PROJ", "summary": "Task 1"},
                    {"project.key": "", "summary": ""},  # Empty row
                    {"project.key": "PROJ", "summary": "Task 2"},
                ]
            )
        )

        # Use mock to simulate file reading
        with patch("builtins.open", mock_open(read_data=csv_content)):
            with patch("os.path.exists", return_value=True):
                with patch("csv.DictReader", return_value=mock_reader):
                    with patch("csv.Sniffer.sniff") as mock_sniffer:
                        # Configure mock for CSV dialect
                        mock_dialect = csv.excel()
                        mock_dialect.delimiter = ","
                        mock_sniffer.return_value = mock_dialect

                        loader = CsvFileLoader()
                        issues = loader.load_issues("test.csv")

                        # Check that empty rows were skipped
                        assert len(issues) == 2
                        assert issues[0]["project"]["key"] == "PROJ"
                        assert issues[0]["summary"] == "Task 1"
                        assert issues[1]["project"]["key"] == "PROJ"
                        assert issues[1]["summary"] == "Task 2"

    def test_csv_nested_dict_conversion(self):
        """Test: conversion of nested dictionaries with dot notation."""
        # Prepare test CSV data with nested keys
        csv_content = """project.key,issuetype.name,assignee.emailAddress,simple_key
PROJ,Task,user@example.com,value"""

        # Create mock for DictReader
        mock_reader = MagicMock()
        mock_reader.fieldnames = [
            "project.key",
            "issuetype.name",
            "assignee.emailAddress",
            "simple_key",
        ]
        mock_reader.__iter__ = Mock(
            return_value=iter(
                [
                    {
                        "project.key": "PROJ",
                        "issuetype.name": "Task",
                        "assignee.emailAddress": "user@example.com",
                        "simple_key": "value",
                    }
                ]
            )
        )

        # Use mock to simulate file reading
        with patch("builtins.open", mock_open(read_data=csv_content)):
            with patch("os.path.exists", return_value=True):
                with patch("csv.DictReader", return_value=mock_reader):
                    with patch("csv.Sniffer.sniff") as mock_sniffer:
                        # Configure mock for CSV dialect
                        mock_dialect = csv.excel()
                        mock_dialect.delimiter = ","
                        mock_sniffer.return_value = mock_dialect

                        loader = CsvFileLoader()
                        issues = loader.load_issues("test.csv")

                        assert len(issues) == 1
                        assert issues[0]["project"]["key"] == "PROJ"
                        assert issues[0]["issuetype"]["name"] == "Task"
                        assert (
                            issues[0]["assignee"]["emailAddress"] == "user@example.com"
                        )
                        assert issues[0]["simple_key"] == "value"

    def test_csv_invalid_format(self):
        """Test: handling invalid CSV error"""
        with patch("builtins.open", mock_open(read_data="col1,col2\nval1,val2")):
            with patch("os.path.exists", return_value=True):
                with patch("csv.Sniffer.sniff") as mock_sniffer:
                    # Configure mock for CSV dialect
                    mock_dialect = csv.excel()
                    mock_dialect.delimiter = ","
                    mock_sniffer.return_value = mock_dialect

                    # Mock DictReader to throw csv.Error during iteration
                    mock_reader = MagicMock()
                    mock_reader.fieldnames = ["col1", "col2"]
                    mock_reader.__iter__.side_effect = csv.Error("Invalid CSV format")

                    with patch("csv.DictReader", return_value=mock_reader):
                        loader = CsvFileLoader()
                        with pytest.raises(ValueError, match="Error parsing CSV file"):
                            loader.load_issues("test.csv")

    def test_excel_empty_file(self):
        """Test: handling empty Excel file error"""
        # Create mock for empty workbook
        mock_sheet = MagicMock()
        mock_sheet.max_row = 0
        mock_sheet.max_column = 0

        mock_workbook = MagicMock()
        mock_workbook.active = mock_sheet
        mock_workbook.close = MagicMock()

        with patch("os.path.exists", return_value=True):
            with patch(
                "tirajira.file_loaders.excel_loader.load_workbook",
                return_value=mock_workbook,
            ):
                with patch.object(ExcelFileLoader, "_validate_and_open_file"):
                    loader = ExcelFileLoader()
                    with pytest.raises(
                        ValueError, match="Excel file is empty or corrupted"
                    ):
                        loader.load_issues("test.xlsx")

    def test_excel_missing_headers(self):
        """Test: handling Excel file with missing headers"""
        # Create mock for workbook with None headers
        mock_sheet = MagicMock()
        mock_sheet.max_row = 2
        mock_sheet.max_column = 2

        # Configure mock for cells with None headers
        mock_sheet.cell.return_value.value = None
        mock_sheet.cell.side_effect = lambda row, column: MagicMock(
            value=None if row == 1 else f"value_{row}_{column}"
        )

        mock_workbook = MagicMock()
        mock_workbook.active = mock_sheet
        mock_workbook.close = MagicMock()

        with patch("os.path.exists", return_value=True):
            with patch(
                "tirajira.file_loaders.excel_loader.load_workbook",
                return_value=mock_workbook,
            ):
                with patch.object(ExcelFileLoader, "_validate_and_open_file"):
                    loader = ExcelFileLoader()
                    issues = loader.load_issues("test.xlsx")

                # Check that data loaded with automatically
                # generated headers
                assert len(issues) == 1
                assert issues[0]["column_1"] == "value_2_1"
                assert issues[0]["column_2"] == "value_2_2"

    def test_excel_skip_empty_rows(self):
        """Test: skipping empty rows in Excel file"""
        # Create mock for workbook with empty rows
        mock_sheet = MagicMock()
        mock_sheet.max_row = 4
        mock_sheet.max_column = 2

        # Configure mock for cells with empty rows
        mock_sheet.cell.return_value.value = None
        mock_sheet.cell.side_effect = lambda row, column: MagicMock(
            value={
                (1, 1): "project.key",
                (1, 2): "summary",
                (2, 1): "PROJ",
                (2, 2): "Task 1",
                (3, 1): None,  # Empty row
                (3, 2): None,  # Empty row
                (4, 1): "PROJ",
                (4, 2): "Task 2",
            }.get((row, column), None)
        )

        mock_workbook = MagicMock()
        mock_workbook.active = mock_sheet
        mock_workbook.close = MagicMock()

        with patch("os.path.exists", return_value=True):
            with patch(
                "tirajira.file_loaders.excel_loader.load_workbook",
                return_value=mock_workbook,
            ):
                with patch.object(ExcelFileLoader, "_validate_and_open_file"):
                    loader = ExcelFileLoader()
                    issues = loader.load_issues("test.xlsx")

                # Check that empty rows were skipped
                assert len(issues) == 2
                assert issues[0]["project"]["key"] == "PROJ"
                assert issues[0]["summary"] == "Task 1"
                assert issues[1]["project"]["key"] == "PROJ"
                assert issues[1]["summary"] == "Task 2"

    def test_excel_nested_dict_conversion(self):
        """Test: conversion of nested dictionaries with dot notation."""
        # Create mock for workbook with nested keys
        mock_sheet = MagicMock()
        mock_sheet.max_row = 2
        mock_sheet.max_column = 4

        # Configure mock for cells with nested keys
        mock_sheet.cell.return_value.value = None
        mock_sheet.cell.side_effect = lambda row, column: MagicMock(
            value={
                (1, 1): "project.key",
                (1, 2): "issuetype.name",
                (1, 3): "assignee.emailAddress",
                (1, 4): "simple_key",
                (2, 1): "PROJ",
                (2, 2): "Task",
                (2, 3): "user@example.com",
                (2, 4): "value",
            }.get((row, column), None)
        )

        mock_workbook = MagicMock()
        mock_workbook.active = mock_sheet
        mock_workbook.close = MagicMock()

        with patch("os.path.exists", return_value=True):
            with patch(
                "tirajira.file_loaders.excel_loader.load_workbook",
                return_value=mock_workbook,
            ):
                with patch.object(ExcelFileLoader, "_validate_and_open_file"):
                    loader = ExcelFileLoader()
                    issues = loader.load_issues("test.xlsx")

                # Check that data loaded correctly with nested dictionaries
                assert len(issues) == 1
                assert issues[0]["project"]["key"] == "PROJ"
                assert issues[0]["issuetype"]["name"] == "Task"
                assert issues[0]["assignee"]["emailAddress"] == "user@example.com"
                assert issues[0]["simple_key"] == "value"

    def test_excel_invalid_format(self):
        """Test: handling invalid Excel file error"""
        with patch("os.path.exists", return_value=True):
            with patch(
                "tirajira.file_loaders.excel_loader.load_workbook"
            ) as mock_load_workbook:
                mock_load_workbook.side_effect = InvalidFileException(
                    "Invalid Excel file format"
                )

                with patch.object(ExcelFileLoader, "_validate_and_open_file"):
                    loader = ExcelFileLoader()
                    with pytest.raises(ValueError, match="Error parsing Excel file"):
                        loader.load_issues("test.xlsx")

    def test_xml_invalid_format(self):
        """Test: handling invalid XML error"""
        # Create loader
        loader = XmlFileLoader()

        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data='<?xml version="1.0" encoding="UTF-8"?><invalid>',
        ):
            with patch("os.path.exists", return_value=True):
                # Check that the correct exception is thrown
                with pytest.raises(ValueError):
                    loader.load_issues("test.xml")

    def test_xml_wrong_root_element(self):
        """Test: handling wrong root element error"""
        # Create loader
        loader = XmlFileLoader()

        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data='<?xml version="1.0" encoding="UTF-8"?><root></root>',
        ):
            with patch("os.path.exists", return_value=True):
                # Check that the correct exception is thrown
                with pytest.raises(ValueError):
                    loader.load_issues("test.xml")

    def test_xml_complex_structure(self):
        """Test: loading XML file with complex structure."""
        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data=(
                '<?xml version="1.0" encoding="UTF-8"?><issues>'
                '<issue id="1"><project key="PROJ"/>'
                "<summary>Test issue 1</summary>"
                "<description>Description for test issue 1</description>"
                '<issuetype name="Task"/>'
                '<assignee email="user1@example.com"/>'
                '<priority name="High"/>'
                "<customfields>"
                '<customfield id="10001" name="Business Value">High</customfield>'
                '<customfield id="10002" name="Story Points">5</customfield>'
                "</customfields></issue>"
                '<issue id="2"><project key="PROJ"/>'
                "<summary>Test issue 2</summary>"
                "<description>Description for test issue 2</description>"
                '<issuetype name="Bug"/>'
                '<assignee email="user2@example.com"/>'
                '<priority name="Medium"/>'
                "<customfields>"
                '<customfield id="10001" name="Business Value">Medium</customfield>'
                '<customfield id="10002" name="Story Points">3</customfield>'
                "</customfields></issue></issues>"
            ),
        ):
            with patch("os.path.exists", return_value=True):
                # Create loader
                loader = XmlFileLoader()

                # Load issues
                issues = loader.load_issues("complex.xml")

                # Check results
                assert len(issues) == 2

                # Check first issue
                issue1 = issues[0]
                assert issue1["project"]["key"] == "PROJ"
                assert issue1["summary"] == "Test issue 1"
                assert issue1["description"] == "Description for test issue 1"
                assert issue1["issuetype"]["name"] == "Task"
                assert issue1["assignee"]["emailAddress"] == "user1@example.com"
                assert issue1["priority"]["name"] == "High"
                assert issue1["customfield_10001"] == "High"
                assert issue1["customfield_10002"] == "5"

                # Check second issue
                issue2 = issues[1]
                assert issue2["project"]["key"] == "PROJ"
                assert issue2["summary"] == "Test issue 2"
                assert issue2["description"] == "Description for test issue 2"
                assert issue2["issuetype"]["name"] == "Bug"
                assert issue2["assignee"]["emailAddress"] == "user2@example.com"
                assert issue2["priority"]["name"] == "Medium"
                assert issue2["customfield_10001"] == "Medium"
                assert issue2["customfield_10002"] == "3"

    def test_xml_dot_notation(self):
        """Test: handling dot notation in attributes."""
        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data=(
                '<?xml version="1.0" encoding="UTF-8"?><issues>'
                '<issue><project key="PROJ"/>'
                "<summary>Test issue</summary>"
                "<description>Description</description>"
                '<issuetype name="Task"/>'
                "<customfields>"
                '<customfield id="customfield_10001" name="Business Value">'
                "High</customfield>"
                "</customfields></issue></issues>"
            ),
        ):
            with patch("os.path.exists", return_value=True):
                # Create loader
                loader = XmlFileLoader()

                # Load issues
                issues = loader.load_issues("dot_notation.xml")

                # Check results
                assert len(issues) == 1
                issue = issues[0]
                assert "customfield_10001" in issue
                assert issue["customfield_10001"] == "High"

    def test_xml_merge_dicts(self):
        """Test: merging dictionaries."""
        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data=(
                '<?xml version="1.0" encoding="UTF-8"?><issues>'
                '<issue><project key="PROJ"/>'
                "<summary>Test issue</summary>"
                "<description>Description</description>"
                '<issuetype name="Task"/>'
                "<labels><label>urgent</label><label>backend</label></labels>"
                "</issue></issues>"
            ),
        ):
            with patch("os.path.exists", return_value=True):
                # Create loader
                loader = XmlFileLoader()

                # Load issues
                issues = loader.load_issues("merge_dicts.xml")

                # Check results
                assert len(issues) == 1
                issue = issues[0]
                assert "labels" in issue
                assert isinstance(issue["labels"], list)
                assert "urgent" in issue["labels"]
                assert "backend" in issue["labels"]

    def test_xml_text_content(self):
        """Test: converting an element with text content to a dictionary."""
        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data=(
                '<?xml version="1.0" encoding="UTF-8"?><issues>'
                '<issue><project key="PROJ"/>'
                "<summary>Simple text content</summary>"
                "</issue></issues>"
            ),
        ):
            with patch("os.path.exists", return_value=True):
                # Create loader
                loader = XmlFileLoader()

                # Load issues
                issues = loader.load_issues("text_content.xml")

                # Check results
                assert len(issues) == 1
                issue = issues[0]
                assert issue["summary"] == "Simple text content"

    def test_xml_mixed_content(self):
        """Test: converting an element with mixed content to a dictionary."""
        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data=(
                '<?xml version="1.0" encoding="UTF-8"?><issues>'
                '<issue><project key="PROJ"/>'
                "<summary>Mixed <b>bold</b> and <i>italic</i> content</summary>"
                "</issue></issues>"
            ),
        ):
            with patch("os.path.exists", return_value=True):
                # Create loader
                loader = XmlFileLoader()

                # Load issues
                issues = loader.load_issues("mixed_content.xml")

                # Check results
                assert len(issues) == 1
                issue = issues[0]

                assert "Mixed bold and italic content" in issue["summary"]

    def test_xml_list_elements(self):
        """Test: converting an element with a list of sub-elements to a dictionary."""
        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data=(
                '<?xml version="1.0" encoding="UTF-8"?><issues>'
                '<issue><project key="PROJ"/>'
                "<attachments>"
                '<attachment name="file1.txt"/>'
                '<attachment name="file2.pdf"/>'
                "</attachments></issue></issues>"
            ),
        ):
            with patch("os.path.exists", return_value=True):
                # Create loader
                loader = XmlFileLoader()

                # Load issues
                issues = loader.load_issues("list_elements.xml")

                # Check results
                assert len(issues) == 1
                issue = issues[0]
                assert "attachments" in issue
                assert isinstance(issue["attachments"], list)
                assert len(issue["attachments"]) == 2
                assert issue["attachments"][0]["name"] == "file1.txt"
                assert issue["attachments"][1]["name"] == "file2.pdf"

    def test_xml_custom_fields_processing(self):
        """Test: processing custom fields."""
        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data=(
                '<?xml version="1.0" encoding="UTF-8"?><issues>'
                '<issue><project key="PROJ"/>'
                "<summary>Test with attributes</summary>"
                '<customfield id="10001">Value1</customfield>'
                '<customfield id="10002">Value2</customfield>'
                "</issue></issues>"
            ),
        ):
            with patch("os.path.exists", return_value=True):
                # Create loader
                loader = XmlFileLoader()

                # Load issues
                issues = loader.load_issues("custom_fields.xml")

                # Check results
                assert len(issues) == 1
                issue = issues[0]
                assert "customfield_10001" in issue
                assert issue["customfield_10001"] == "Value1"
                assert "customfield_10002" in issue
                assert issue["customfield_10002"] == "Value2"

    def test_xml_nested_elements_with_attributes(self):
        """Test: processing nested elements with attributes."""
        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data=(
                '<?xml version="1.0" encoding="UTF-8"?><issues>'
                '<issue><project key="PROJ"/>'
                "<summary>Test with nested elements</summary>"
                '<assignee email="test@example.com" name="Test User"/>'
                '<priority name="High"/>'
                "</issue></issues>"
            ),
        ):
            with patch("os.path.exists", return_value=True):
                # Create loader
                loader = XmlFileLoader()

                # Load issues
                issues = loader.load_issues("nested_elements.xml")

                # Check results
                assert len(issues) == 1
                issue = issues[0]
                assert issue["assignee"]["emailAddress"] == "test@example.com"
                assert issue["assignee"]["name"] == "Test User"
                assert issue["priority"]["name"] == "High"

    def test_xml_mixed_content_with_tail(self):
        """Test: processing elements with mixed content and text after
        child elements."""
        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data=(
                '<?xml version="1.0" encoding="UTF-8"?><issues>'
                '<issue><project key="PROJ"/>'
                "<summary>Description<b>bold</b> and <i>italic</i> text</summary>"
                "</issue></issues>"
            ),
        ):
            with patch("os.path.exists", return_value=True):
                # Create loader
                loader = XmlFileLoader()

                # Load issues
                issues = loader.load_issues("mixed_content_tail.xml")

                # Check results
                assert len(issues) == 1
                issue = issues[0]
                # Check that the text content is properly merged
                summary = issue["summary"]
                assert isinstance(summary, str)
                assert "Description" in summary
                assert "bold" in summary
                assert "and" in summary
                assert "italic" in summary
                assert "text" in summary

    def test_xml_empty_and_whitespace_elements(self):
        """Test: processing empty elements and elements with whitespace."""
        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data=(
                '<?xml version="1.0" encoding="UTF-8"?><issues>'
                '<issue><project key="PROJ"/>'
                "<empty_element></empty_element>"
                "<whitespace_element>   </whitespace_element>"
                "</issue></issues>"
            ),
        ):
            with patch("os.path.exists", return_value=True):
                # Create loader
                loader = XmlFileLoader()

                # Load issues
                issues = loader.load_issues("empty_whitespace.xml")

                # Check results
                assert len(issues) == 1
                issue = issues[0]
                # Empty elements should be represented as empty strings
                assert issue["empty_element"] == ""
                # Elements with whitespace should be cleaned
                assert issue["whitespace_element"] == ""

    def test_xml_dot_notation_in_element_names(self):
        """Test: handling dot notation in element names."""
        with patch(
            "builtins.open",
            new_callable=mock_open,
            read_data=(
                '<?xml version="1.0" encoding="UTF-8"?><issues>'
                "<issue><project.key>PROJ</project.key>"
                "<summary>Dot notation in element names</summary>"
                "</issue></issues>"
            ),
        ):
            with patch("os.path.exists", return_value=True):
                # Create loader
                loader = XmlFileLoader()

                # Load issues
                issues = loader.load_issues("dot_notation_elements.xml")

                # Check results
                assert len(issues) == 1
                issue = issues[0]

                assert "project" in issue
                assert isinstance(issue["project"], dict)
                assert "key" in issue["project"]
                assert issue["project"]["key"] == "PROJ"

    def test_xml_merge_dicts_logic(self):
        """Test: merging dictionaries logic in _merge_dicts method."""
        # Create loader
        loader = XmlFileLoader()

        # Test merging two simple dictionaries
        target = {"a": 1, "b": 2}
        source = {"b": 3, "c": 4}
        loader._merge_dicts(target, source)
        assert target == {"a": 1, "b": [2, 3], "c": 4}

        # Test merging nested dictionaries
        target = {"a": {"x": 1}, "b": 2}
        source = {"a": {"y": 2}, "c": 3}
        loader._merge_dicts(target, source)
        assert target == {"a": {"x": 1, "y": 2}, "b": 2, "c": 3}

        # Test adding to an existing list
        target = {"a": [1, 2]}
        source = {"a": 3}
        loader._merge_dicts(target, source)
        assert target == {"a": [1, 2, 3]}

    def test_xml_process_list_field_logic(self):
        """Test: processing fields logic in _process_list_field method."""
        # Create loader
        loader = XmlFileLoader()

        # Test processing labels with a single element
        result = {"labels": {"label": "bug"}}
        loader._process_list_field(result, "labels", "label")
        assert result == {"labels": ["bug"]}

        # Test processing labels with multiple elements
        result = {"labels": {"label": ["bug", "frontend"]}}
        loader._process_list_field(result, "labels", "label")
        assert result == {"labels": ["bug", "frontend"]}

        # Test processing labels with a missing key
        result = {"labels": {"other": "bug"}}
        loader._process_list_field(result, "labels", "label")
        assert result == {"labels": {"other": "bug"}}

        # Test processing a missing field
        result = {"other": "value"}
        loader._process_list_field(result, "labels", "label")
        assert result == {"other": "value"}

    def test_xml_process_labels_logic(self):
        """Test: processing labels logic in _process_labels method."""
        # Create loader
        loader = XmlFileLoader()

        # Test processing labels
        result = {"labels": {"label": "bug"}}
        loader._process_labels(result)
        assert result == {"labels": ["bug"]}

    def test_xml_process_attachments_logic(self):
        """Test: processing attachments logic in _process_attachments method."""
        # Create loader
        loader = XmlFileLoader()

        # Test processing attachments with a single element
        result = {"attachments": {"attachment": {"name": "file.txt"}}}
        loader._process_attachments(result)
        assert result == {"attachments": [{"name": "file.txt"}]}

    def test_xml_process_customfields_post_logic(self):
        """Test: processing customfields logic in _process_customfields_post method."""
        # Create loader
        loader = XmlFileLoader()

        # Test processing customfields
        result = {
            "customfields": {
                "customfield_10001": "Value1",
                "customfield_10002": "Value2",
                "other_field": "Other",
            },
            "summary": "Test",
        }
        expected = {
            "customfield_10001": "Value1",
            "customfield_10002": "Value2",
            "summary": "Test",
        }
        loader._process_customfields_post(result)
        assert result == expected

        # Test processing missing customfields
        result = {"summary": "Test"}
        expected = {"summary": "Test"}
        loader._process_customfields_post(result)
        assert result == expected

    def test_xml_process_attributes_logic(self):
        """Test _process_attributes method logic for handling XML element attributes."""
        loader = XmlFileLoader()

        # Test special attributes conversion
        import xml.etree.ElementTree as ET

        element = ET.Element("test")
        element.set("email", "user@example.com")
        element.set("name", "Test User")
        element.set("key", "TEST-123")

        result = {}
        loader._process_attributes(element, result)

        assert result["emailAddress"] == "user@example.com"
        assert result["name"] == "Test User"
        assert result["key"] == "TEST-123"

        # Test regular attributes preservation
        element = ET.Element("test")
        element.set("priority", "High")
        element.set("status", "Open")

        result = {}
        loader._process_attributes(element, result)

        assert result["priority"] == "High"
        assert result["status"] == "Open"

        # Test dot notation attribute handling
        element = ET.Element("test")
        element.set("project.key", "PROJ")

        result = {}
        loader._process_attributes(element, result)

        # Verify dot notation was processed
        assert "project" in result or "project.key" in result

    def test_xml_process_children_logic(self):
        """Test _process_children method logic for handling child elements."""
        loader = XmlFileLoader()

        # Test processing single child element
        import xml.etree.ElementTree as ET

        parent = ET.Element("parent")
        child = ET.SubElement(parent, "summary")
        child.text = "Test summary"

        result = loader._process_children(parent)

        assert "summary" in result
        assert result["summary"] == "Test summary"

        # Test processing multiple child elements with same tag (should create list)
        parent = ET.Element("parent")
        child1 = ET.SubElement(parent, "label")
        child1.text = "bug"
        child2 = ET.SubElement(parent, "label")
        child2.text = "frontend"

        result = loader._process_children(parent)

        assert "label" in result
        assert isinstance(result["label"], list)
        assert len(result["label"]) == 2
        assert "bug" in result["label"]
        assert "frontend" in result["label"]

        # Test processing customfield elements (special handling)
        parent = ET.Element("parent")
        child1 = ET.SubElement(parent, "customfield", {"id": "10001"})
        child1.text = "Value1"
        child2 = ET.SubElement(parent, "customfield", {"id": "10002"})
        child2.text = "Value2"

        result = loader._process_children(parent)

        assert "customfield_10001" in result
        assert result["customfield_10001"] == "Value1"
        assert "customfield_10002" in result
        assert result["customfield_10002"] == "Value2"

    def test_xml_process_customfields_logic(self):
        """Test _process_customfields method logic for handling
        custom field elements."""
        loader = XmlFileLoader()

        # Test processing customfield elements with numeric ids
        import xml.etree.ElementTree as ET

        children = []
        child1 = ET.Element("customfield", {"id": "10001"})
        child1.text = "Value1"
        children.append(child1)

        child2 = ET.Element("customfield", {"id": "10002"})
        child2.text = "Value2"
        children.append(child2)

        children_dict = {}
        loader._process_customfields(children, children_dict)

        assert "customfield_10001" in children_dict
        assert children_dict["customfield_10001"] == "Value1"
        assert "customfield_10002" in children_dict
        assert children_dict["customfield_10002"] == "Value2"

        # Test processing customfield elements with existing customfield_ prefix
        children = []
        child = ET.Element("customfield", {"id": "customfield_10003"})
        child.text = "Value3"
        children.append(child)

        children_dict = {}
        loader._process_customfields(children, children_dict)

        assert "customfield_10003" in children_dict
        assert children_dict["customfield_10003"] == "Value3"

        # Test processing customfield elements with empty text
        children = []
        child = ET.Element("customfield", {"id": "10004"})
        child.text = None
        children.append(child)

        children_dict = {}
        loader._process_customfields(children, children_dict)

        assert "customfield_10004" in children_dict
        assert children_dict["customfield_10004"] == ""

    def test_xml_process_child_element_logic(self):
        """Test _process_child_element method logic for handling
        individual child elements."""
        loader = XmlFileLoader()

        # Test processing simple text element
        import xml.etree.ElementTree as ET

        child = ET.Element("summary")
        child.text = "Test summary"

        result = loader._process_child_element(child)

        assert result == "Test summary"

        # Test processing element with tail text
        child = ET.Element("summary")
        child.text = "Text "
        bold = ET.SubElement(child, "b")
        bold.text = "bold"
        bold.tail = " text"

        result = loader._process_child_element(child)

        assert "Text bold text" in result

        # Test processing element with children (recursive processing)
        child = ET.Element("project")
        key = ET.SubElement(child, "key")
        key.text = "PROJ"

        result = loader._process_child_element(child)

        assert isinstance(result, dict)
        assert result["key"] == "PROJ"

        # Test processing empty element
        child = ET.Element("empty")
        child.text = None

        result = loader._process_child_element(child)

        assert result == ""

    def test_xml_add_to_children_dict_logic(self):
        """Test _add_to_children_dict method logic for adding values
        to children dictionary."""
        loader = XmlFileLoader()

        # Test adding simple key-value pair
        children_dict = {}
        loader._add_to_children_dict("priority", "High", children_dict)

        assert "priority" in children_dict
        assert children_dict["priority"] == "High"

        # Test adding key with dot notation (should be converted to nested dict)
        children_dict = {}
        loader._add_to_children_dict("project.key", "PROJ", children_dict)

        assert "project" in children_dict or "project.key" in children_dict

    def test_xml_post_process_result_logic(self):
        """Test _post_process_result method logic for post-processing results."""
        loader = XmlFileLoader()

        # Test labels post-processing
        result = {"labels": {"label": "bug"}}
        loader._post_process_result(result)

        assert "labels" in result
        assert isinstance(result["labels"], list)
        assert "bug" in result["labels"]

        # Test attachments post-processing
        result = {"attachments": {"attachment": {"name": "file.txt"}}}
        loader._post_process_result(result)

        assert "attachments" in result
        assert isinstance(result["attachments"], list)
        assert len(result["attachments"]) == 1
        assert result["attachments"][0]["name"] == "file.txt"

        # Test customfields post-processing
        result = {
            "customfields": {
                "customfield_10001": "Value1",
                "customfield_10002": "Value2",
            },
            "summary": "Test",
        }
        loader._post_process_result(result)

        # Custom fields should be moved to top level
        assert "customfield_10001" in result
        assert result["customfield_10001"] == "Value1"
        assert "customfield_10002" in result
        assert result["customfield_10002"] == "Value2"
        assert "customfields" not in result
        assert result["summary"] == "Test"

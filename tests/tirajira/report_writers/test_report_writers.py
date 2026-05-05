"""
Tests for all report writers using parametrized tests.
"""

import os
import tempfile
import xml.etree.ElementTree as ET
from unittest.mock import mock_open, patch

import pytest
from openpyxl import load_workbook

from tests.tirajira.test_data import COMMON_REPORT_DATA
from tirajira.report_writers.csv_writer import CsvReportWriter
from tirajira.report_writers.excel_writer import ExcelReportWriter
from tirajira.report_writers.factory import create_report_writer
from tirajira.report_writers.json_writer import JsonReportWriter
from tirajira.report_writers.xml_writer import XmlReportWriter
from tirajira.report_writers.yaml_writer import YamlReportWriter


@pytest.mark.parametrize(
    "writer_class,file_extension,expected_calls_kwargs",
    [
        (JsonReportWriter, "report.json", {"encoding": "utf-8"}),
        (CsvReportWriter, "report.csv", {"newline": "", "encoding": "utf-8"}),
        (YamlReportWriter, "report.yaml", {"encoding": "utf-8"}),
    ],
)
@patch("builtins.open", new_callable=mock_open)
def test_text_based_report_writers_write_report(
    mock_file, writer_class, file_extension, expected_calls_kwargs
):
    """Test: writing report for text formats (JSON, CSV, YAML)"""
    writer = writer_class()
    writer.write_report(COMMON_REPORT_DATA, file_extension)
    mock_file.assert_called_once_with(file_extension, "w", **expected_calls_kwargs)


@pytest.mark.parametrize(
    "format_name,expected_class",
    [
        ("json", JsonReportWriter),
        ("yaml", YamlReportWriter),
        ("yml", YamlReportWriter),
        ("csv", CsvReportWriter),
        ("excel", ExcelReportWriter),
        ("xlsx", ExcelReportWriter),
        ("xml", XmlReportWriter),
    ],
)
def test_create_report_writer(format_name, expected_class):
    """Test: creating report writer for all supported formats"""
    writer = create_report_writer(format_name)
    assert isinstance(writer, expected_class)


def test_excel_report_writer_write_report():
    """Test: writing report in Excel format"""
    writer = ExcelReportWriter()
    report_data = COMMON_REPORT_DATA.copy()
    report_data["tasks"][0]["issue_data"]["description"] = "Description of test task"
    report_data["tasks"][0]["assignee"] = {
        "name": "John Doe",
        "email": "john.doe@example.com",
    }
    report_data["tasks"][0]["labels"] = ["urgent", "backend"]

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        tmp_file_path = tmp_file.name

    try:
        writer.write_report(report_data, tmp_file_path)
        assert os.path.exists(tmp_file_path)

        workbook = load_workbook(tmp_file_path)
        assert "Metadata" in workbook.sheetnames
        assert "Tasks" in workbook.sheetnames

        metadata_sheet = workbook["Metadata"]
        assert metadata_sheet["A1"].value == "generated_at"
        assert metadata_sheet["B1"].value == "2023-12-01T15:30:45"

        tasks_sheet = workbook["Tasks"]
        assert tasks_sheet["A1"].value == "assignee.email"
        assert tasks_sheet["B1"].value == "assignee.name"

    finally:
        if os.path.exists(tmp_file_path):
            os.unlink(tmp_file_path)


def test_excel_report_writer_empty_data():
    """Test: writing report with empty data in Excel format"""
    writer = ExcelReportWriter()
    report_data = {
        "metadata": {
            "generated_at": "2023-12-01T15:30:45",
            "source_file": "test.json",
            "total_tasks": 0,
            "successful_tasks": 0,
            "failed_tasks": 0,
        },
        "tasks": [],
    }

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        tmp_file_path = tmp_file.name

    try:
        writer.write_report(report_data, tmp_file_path)
        assert os.path.exists(tmp_file_path)

        workbook = load_workbook(tmp_file_path)
        assert "Metadata" in workbook.sheetnames
        assert "Tasks" in workbook.sheetnames

    finally:
        if os.path.exists(tmp_file_path):
            os.unlink(tmp_file_path)


def test_xml_report_writer_write_report():
    """Test: writing report in XML format"""
    writer = XmlReportWriter()
    report_data = COMMON_REPORT_DATA

    with tempfile.NamedTemporaryFile(suffix=".xml", delete=False) as tmp_file:
        tmp_file_path = tmp_file.name

    try:
        writer.write_report(report_data, tmp_file_path)
        assert os.path.exists(tmp_file_path)

        tree = ET.parse(tmp_file_path)
        root = tree.getroot()

        assert root.tag == "report"

        metadata_element = root.find("metadata")
        tasks_element = root.find("tasks")
        assert metadata_element is not None
        assert tasks_element is not None

        generated_at_element = metadata_element.find("generated_at")
        assert generated_at_element is not None
        assert generated_at_element.text == "2023-12-01T15:30:45"

        task_elements = tasks_element.findall("task")
        assert len(task_elements) == 2

        first_task = task_elements[0]
        status_element = first_task.find("status")
        assert status_element is not None
        assert status_element.text == "success"

    finally:
        if os.path.exists(tmp_file_path):
            os.unlink(tmp_file_path)


def test_xml_report_writer_with_nested_data():
    """Test: writing report with nested data in XML format"""
    writer = XmlReportWriter()
    report_data = {
        "metadata": {
            "generated_at": "2023-12-01T15:30:45",
            "source_file": "nested_test.xml",
            "total_tasks": 1,
            "successful_tasks": 1,
            "failed_tasks": 0,
        },
        "tasks": [
            {
                "id": 0,
                "status": "success",
                "issue_key": "PROJ-124",
                "issue_data": {
                    "summary": "Nested data task",
                    "description": "Task with nested data structures",
                    "assignee": {"name": "John Doe", "email": "john.doe@example.com"},
                    "custom_fields": {"business_value": "High", "story_points": 5},
                },
                "labels": ["urgent", "backend"],
                "processed_at": "2023-12-01T15:30:47",
            }
        ],
    }

    with tempfile.NamedTemporaryFile(suffix=".xml", delete=False) as tmp_file:
        tmp_file_path = tmp_file.name

    try:
        writer.write_report(report_data, tmp_file_path)
        assert os.path.exists(tmp_file_path)

        tree = ET.parse(tmp_file_path)
        root = tree.getroot()

        assert root.tag == "report"

        metadata_element = root.find("metadata")
        tasks_element = root.find("tasks")
        assert metadata_element is not None
        assert tasks_element is not None

        task_elements = tasks_element.findall("task")
        assert len(task_elements) == 1

        task = task_elements[0]
        issue_data_element = task.find("issue_data")
        assert issue_data_element is not None

        assignee_element = issue_data_element.find("assignee")
        assert assignee_element is not None

        name_element = assignee_element.find("name")
        assert name_element is not None
        assert name_element.text == "John Doe"

    finally:
        if os.path.exists(tmp_file_path):
            os.unlink(tmp_file_path)

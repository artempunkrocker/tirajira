"""
Тесты для Excel писателя отчетов.
"""

import os
import tempfile

from openpyxl import load_workbook

from tirajira.report_writers.excel_writer import ExcelReportWriter


def test_excel_report_writer_write_report():
    """Тест: запись отчета в формате Excel"""
    # Создаем писатель отчетов
    writer = ExcelReportWriter()

    # Создаем данные отчета с вложенными структурами и списками
    report_data = {
        "metadata": {
            "generated_at": "2023-12-01T15:30:45",
            "source_file": "test.json",
            "total_tasks": 2,
            "successful_tasks": 1,
            "failed_tasks": 1,
        },
        "tasks": [
            {
                "id": 0,
                "status": "success",
                "issue_key": "PROJ-123",
                "issue_data": {
                    "summary": "Test task",
                    "description": "Description of test task",
                },
                "assignee": {"name": "John Doe", "email": "john.doe@example.com"},
                "labels": ["urgent", "backend"],
                "processed_at": "2023-12-01T15:30:45",
            },
            {
                "id": 1,
                "status": "failure",
                "error_message": "Connection error",
                "issue_data": {
                    "summary": "Another task",
                    "description": "Description of another task",
                },
                "processed_at": "2023-12-01T15:30:46",
            },
        ],
    }

    # Создаем временный файл для теста
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        tmp_file_path = tmp_file.name

    try:
        # Записываем отчет
        writer.write_report(report_data, tmp_file_path)

        # Проверяем, что файл был создан
        assert os.path.exists(tmp_file_path)

        # Проверяем содержимое файла
        workbook = load_workbook(tmp_file_path)
        assert "Metadata" in workbook.sheetnames
        assert "Tasks" in workbook.sheetnames

        # Проверяем содержимое листа Metadata
        metadata_sheet = workbook["Metadata"]
        assert metadata_sheet["A1"].value == "generated_at"
        assert metadata_sheet["B1"].value == "2023-12-01T15:30:45"

        # Проверяем содержимое листа Tasks
        tasks_sheet = workbook["Tasks"]
        assert tasks_sheet["A1"].value == "assignee.email"
        assert tasks_sheet["B1"].value == "assignee.name"

    finally:
        # Удаляем временный файл
        if os.path.exists(tmp_file_path):
            os.unlink(tmp_file_path)


def test_excel_report_writer_empty_data():
    """Тест: запись отчета с пустыми данными"""
    # Создаем писатель отчетов
    writer = ExcelReportWriter()

    # Создаем пустые данные отчета
    report_data = {
        "metadata": {
            "generated_at": "2023-12-01T15:30:45",
            "source_file": "test.json",
            "total_tasks": 0,
            "successful_tasks": 0,
            "failed_tasks": 0,
        },
        "tasks": [],
    }

    # Создаем временный файл для теста
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        tmp_file_path = tmp_file.name

    try:
        # Записываем отчет
        writer.write_report(report_data, tmp_file_path)

        # Проверяем, что файл был создан
        assert os.path.exists(tmp_file_path)

        # Проверяем содержимое файла
        workbook = load_workbook(tmp_file_path)
        assert "Metadata" in workbook.sheetnames
        assert "Tasks" in workbook.sheetnames

    finally:
        # Удаляем временный файл
        if os.path.exists(tmp_file_path):
            os.unlink(tmp_file_path)
